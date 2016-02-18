import openpyxl
import shutil
import os
import glob

# _________Notes_________
# Written by Cody Towstik, 2016
# File must have group # in it
# 'final peer evaluation.xlsx' file must be in same directory as this file and the excel files
# There must be no backslash when inputting the directory to be worked on


# Makes a list of all of the excel files in the directory and finds how many there are
fileList = glob.glob('*.xlsx')
numOfFiles = len(fileList)

# ##New directory creation info###
srcfile = input('What is the folder directory? (Make sure there is no backslash at the end) \n ::-> ') + '\\final peer evaluation.xlsx'
# srcfile = "C:\\users\cody\\pycharmprojects\\exceltowtie\\final peer evaluation.xlsx"

newPath1 = input('Where do you want to put them (directory)? \n ::-> ')
# newPath1 = 'C:\\Users\Cody\\PycharmProjects\\ExcelTowtie\\FileDump'

if not os.path.exists(newPath1):  # Checks if the folder is already made, then makes it.
    os.makedirs(newPath1)

# Name of group members to start looking for
groupNames = []

numberOfGroups = 3  # starts checking for group number

for groupNumber in range(1, numberOfGroups+1):
    # progress notification
    print('Starting group ', groupNumber, '...', sep='')
    # loops first through each file to populate group names, could be optimized but eh
    for file in fileList:
        if str(groupNumber) in file:
            # opens workbook and active sheet
            wb = openpyxl.load_workbook(file)
            sheet = wb.active

            # reads in names
            for x in sheet.rows[12]:
                # Checks if there is a name
                if x.value:
                    groupNames.append(x.value)

    # makes the list unique to remove copies, semi robust name check :)
    groupNames = list(set(groupNames))

    groupFolder = newPath1+'\\'+str(groupNumber)
    # make the directory for that group
    if not os.path.exists(groupFolder):  # Checks if the folder is already made, then makes it.
        os.makedirs(groupFolder)

    # Starts with each name to search for
    for name in groupNames:
        # Keeps track of position in new document to find and paste comments
        columnCounter = 0
        columnList = ['B', 'C', 'D', 'E', 'F', 'G']
        rowCounter = 12

        # copy template and rename to Last_First_Group#_Completed
        groupTempDir = groupFolder+'\\'+'final peer evaluation.xlsx'
        if not os.path.exists(groupTempDir):
            shutil.copy(srcfile, groupFolder)

        copyDir = groupFolder+'\\'+name+' Peer Eval.xlsx'
        if not os.path.exists(copyDir):
            os.rename(groupTempDir, copyDir)

        # Creates the workbook to be copied to from all of the individual files
        wbCopy = openpyxl.load_workbook(copyDir)
        sheetCopy = wbCopy.active
        sheetCopy.title = name

        # looks for each file in that group to find the name
        for file in fileList:
            # works the copy and paste magic
            if str(groupNumber) in file:
                # opens workbook and active sheet
                wb = openpyxl.load_workbook(file)
                sheet = wb.active

                # Keeps track of where the name was found
                inputColCounter = 0

                # reads in names
                for xy in sheet.rows[rowCounter]:  # optimize by breaking for loop after name is found
                    # Checks if the name is the current one
                    inputColCounter += 1

                    if xy.value == name:
                        while rowCounter < 45:
                            # check if it is their own file
                            if not sheet.cell(row=19, column=inputColCounter).value and not sheet.cell(row=30, column=inputColCounter).value:
                                columnCounter -= 1
                                break
                            inputCell = sheet.cell(row=rowCounter+1, column=inputColCounter)
                            #copies cell data
                            if inputCell.value:
                                sheetCopy['{0}{1}'.format(columnList[columnCounter], rowCounter+1)].value = inputCell.value
                            rowCounter += 1
                        rowCounter = 12
                        columnCounter += 1
                        break

                wbCopy.save(copyDir)
        wbCopy.save(copyDir)
        # resets column counter after that persons file is filled
        columnCounter = 0
    # resets groupName List
    groupNames = []

